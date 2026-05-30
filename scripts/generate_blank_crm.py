#!/usr/bin/env python3
"""
generate_blank_crm.py

Creates a blank SmartCRM template xlsx with all tabs, headers, dropdowns,
and formatting — but ZERO contact data.

For distribution as a blank starting point. No personal data included.

Usage:
    python3 scripts/generate_blank_crm.py

Output:
    templates/SmartCRM_Blank_Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Colors ────────────────────────────────────────────────
NAVY_HEX     = "0D1B2A"
NAVY_MID_HEX = "1A2F45"
GOLD_HEX     = "C9A84C"
WHITE_HEX    = "F4F1EB"
TEAL_HEX     = "2EB8B8"

HEADER_FILL  = PatternFill("solid", start_color=NAVY_HEX)
HEADER_FONT  = Font(name="Arial", bold=True, color=WHITE_HEX, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT   = Font(name="Arial", bold=True, color=NAVY_HEX, size=16)
BODY_FONT    = Font(name="Arial", size=10)

STATUS_VALUES = (
    '"1. To Connect,2. Request Sent,3. Request Accepted,4. Intro Sent,'
    '5. Meeting Requested,6. Scheduled,7. Replied — Positive,'
    '8. Replied — Not Now,9. Follow-up 1 (4d),10. Follow-up 2 (10d),'
    '11. Not a Fit,12. Closed — Won,13. Closed — Lost"'
)
PRIORITY_VALUES = '"Tier 0,Tier 1,Tier 2,High,Med,Low"'

# ─── Silo schemas ──────────────────────────────────────────
SILO_SCHEMAS = {
    "Venture & Investment": [
        "Priority", "Status", "Next Action Date", "Name", "Title / Role",
        "Company / Fund", "LinkedIn URL", "X Handle",
        "Context / Relationship", "Connection Note", "Intro Message",
        "Meeting Ask", "Last Touch Date", "Variant ID",
        "Reply Snippet", "Enrichment", "Notes & Research"
    ],
    "OTC & Market Makers": [
        "Priority", "Status", "Next Action Date", "Name", "Title / Role",
        "Company", "LinkedIn URL", "X Handle",
        "Context / Relationship", "Connection Note", "Intro Message",
        "Meeting Ask", "Last Touch Date", "Variant ID",
        "Reply Snippet", "Enrichment", "Notes & Research"
    ],
    "Partnerships": [
        "Priority", "Status", "Next Action Date", "Name", "Title / Role",
        "Company / Institution", "LinkedIn URL", "X Handle",
        "Context / Relationship", "Connection Note", "Intro Message",
        "Meeting Ask", "Last Touch Date", "Variant ID",
        "Reply Snippet", "Enrichment", "Notes & Research"
    ],
    "Freight Forwarders": [
        "Priority", "Status", "Next Action Date", "Name", "Title / Role",
        "Company", "LinkedIn URL", "X Handle",
        "Context / Relationship", "Connection Note", "Intro Message",
        "Meeting Ask", "Last Touch Date", "Variant ID",
        "Reply Snippet", "Enrichment", "Notes & Research"
    ],
    "General Contacts": [
        "Priority", "Status", "Next Action Date", "Name", "Title / Role",
        "Company", "LinkedIn URL", "X Handle",
        "Context / Relationship", "Connection Note", "Intro Message",
        "Meeting Ask", "Last Touch Date", "Variant ID",
        "Reply Snippet", "Enrichment", "Notes & Research"
    ],
}

COL_WIDTHS = {
    "Priority": 12, "Status": 24, "Next Action Date": 16,
    "Name": 22, "Title / Role": 22, "Company / Fund": 22,
    "Company / Institution": 22, "Company": 22,
    "LinkedIn URL": 26, "X Handle": 16,
    "Context / Relationship": 24, "Connection Note": 36,
    "Intro Message": 36, "Meeting Ask": 30,
    "Last Touch Date": 16, "Variant ID": 14,
    "Reply Snippet": 28, "Enrichment": 40,
    "Notes & Research": 50,
}


def add_silo(wb, name, headers):
    ws = wb.create_sheet(name)

    # Row 1: Title
    ws["A1"] = name
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws.row_dimensions[1].height = 28

    # Row 2: Headers
    ws.row_dimensions[2].height = 36
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(header, 18)

    # Data validation dropdowns (rows 3–502)
    dv_status = DataValidation(
        type="list", formula1=STATUS_VALUES, allow_blank=True,
        showDropDown=False
    )
    dv_status.sqref = "B3:B502"
    ws.add_data_validation(dv_status)

    dv_priority = DataValidation(
        type="list", formula1=PRIORITY_VALUES, allow_blank=True,
        showDropDown=False
    )
    dv_priority.sqref = "A3:A502"
    ws.add_data_validation(dv_priority)

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Leave 5 example rows with placeholder text (shows structure, no real data)
    for row in range(3, 6):
        ws.cell(row=row, column=1, value="Tier 1")
        ws.cell(row=row, column=2, value="1. To Connect")
        ws.cell(row=row, column=4, value=f"[Contact Name {row-2}]")
        ws.cell(row=row, column=5, value="[Title]")
        ws.cell(row=row, column=6, value="[Company]")
        ws.cell(row=row, column=7, value="[LinkedIn URL]")

    return ws


def add_grants_tab(wb):
    ws = wb.create_sheet("🎯 Grants Pipeline")
    headers = [
        "Priority", "Status", "Deadline", "Date Applied", "Grant Name",
        "Category", "Ecosystem", "Est. Size", "Application URL",
        "Contact Email/Social", "Pitch Angle", "Amount Requested",
        "Amount Awarded", "Next Action"
    ]
    ws["A1"] = "Grants Pipeline"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    ws.freeze_panes = "A3"
    col_widths = {1:10, 2:16, 3:14, 4:14, 5:32, 6:18, 7:18, 8:12,
                  9:32, 10:24, 11:40, 12:16, 13:16, 14:28}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def add_drafts_log(wb):
    ws = wb.create_sheet("📨 Drafts Log")
    headers = [
        "Date", "Contact Name", "Silo", "Platform", "Touch Type",
        "Variant ID", "Draft Text", "Outcome", "Outcome Date"
    ]
    ws["A1"] = "Drafts Log"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    col_widths = {1:12, 2:22, 3:22, 4:12, 5:24, 6:14, 7:60, 8:24, 9:14}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"


def add_readme_tab(wb):
    ws = wb.create_sheet("📖 README", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    ws["A1"] = "SmartCRM — Blank Template"
    ws["A1"].font = Font(name="Arial", bold=True, size=20, color=NAVY_HEX)

    ws["A2"] = "Fill in project_facts.md and founder_voice.md in your bot's context/ folder before running the bot."
    ws["A2"].font = Font(name="Arial", italic=True, color="666666", size=10)

    lines = [
        ("", ""),
        ("TABS", ""),
        ("📊 Dashboard", "Add charts here after import (Insert → Chart on data ranges)"),
        ("🎯 Grants Pipeline", "Add grants from the grant-application skill documentation"),
        ("Venture & Investment", "VC contacts"),
        ("OTC & Market Makers", "Liquidity desk contacts"),
        ("Partnerships", "Strategic partnerships"),
        ("Freight Forwarders", "Logistics contacts"),
        ("General Contacts", "Your full network"),
        ("📨 Drafts Log", "Every draft the bot generates"),
        ("", ""),
        ("SETUP", ""),
        ("1. Import to Google Sheets", "File → Import → this xlsx → replace spreadsheet"),
        ("2. Rename to 'SmartCRM'", "Click the title in the top-left"),
        ("3. Get the Sheet ID", "From the URL: /d/[THIS PART]/edit"),
        ("4. Follow SETUP.md", "Complete the full setup guide"),
        ("", ""),
        ("STATUS LIFECYCLE", ""),
        ("1. To Connect", "Identified, not contacted yet"),
        ("2. Request Sent", "Connection request sent — bot sets Next Action Date +7 days"),
        ("3. Request Accepted", "Connected — bot enriches and drafts intro DM"),
        ("4. Intro Sent", "First substantive DM sent"),
        ("5. Meeting Requested", "Asked for a call"),
        ("6. Scheduled", "Meeting on calendar"),
        ("7. Replied — Positive", "Interested, not yet scheduled"),
        ("8. Replied — Not Now", "Soft no, keep warm"),
        ("9. Follow-up 1 (4d)", "4-day follow-up due"),
        ("10. Follow-up 2 (10d)", "10-day follow-up due"),
        ("11. Not a Fit", "Hard no, archive"),
        ("12. Closed — Won", "Deal/pilot signed"),
        ("13. Closed — Lost", "Dropped after engagement"),
    ]

    for i, (a, b) in enumerate(lines, start=4):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        if a and not b:
            ws.cell(row=i, column=1).font = Font(name="Arial", bold=True, size=13, color=NAVY_HEX)
        elif a:
            ws.cell(row=i, column=1).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=i, column=2).font = Font(name="Arial", size=10)


def main():
    os.makedirs("templates", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    add_readme_tab(wb)

    # Placeholder Dashboard tab
    dash = wb.create_sheet("📊 Dashboard")
    dash["A1"] = "Dashboard"
    dash["A1"].font = TITLE_FONT
    dash["A2"] = "Add charts here pointing at silo data once the Sheet is in Google Sheets."
    dash["A2"].font = Font(name="Arial", size=10, color="666666")

    for silo_name, headers in SILO_SCHEMAS.items():
        add_silo(wb, silo_name, headers)

    add_grants_tab(wb)
    add_drafts_log(wb)

    out_path = "templates/SmartCRM_Blank_Template.xlsx"
    wb.save(out_path)
    print(f"✓ Blank template saved to {out_path}")
    print("  No contact data included. Safe to distribute.")
    print("")
    print("  To use: import this file into Google Sheets,")
    print("  then follow SETUP.md to connect it to the bot.")


if __name__ == "__main__":
    main()
